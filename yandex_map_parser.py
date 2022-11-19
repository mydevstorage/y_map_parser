import os
import sub
import sys
import warnings
from time import sleep
from loguru import logger
from threading import Thread
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium_stealth import stealth
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import load_workbook
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities

warnings.filterwarnings('ignore')

URL = "https://yandex.ru/maps"

useragent = ("user-agent=Mozilla/5.0 (X11; Linux x86_64) '\
                         'AppleWebKit/537.36 (KHTML, like Gecko) '\
                         'Chrome/106.0.0.0 Safari/537.36")

PATH_TO_DRIVER = ('/home/roman/real_python/web_parsing/'
                  'yandex_map_parser/chromedriver')

logger.remove(0)
logger.add(sys.stderr, format="<green>{time:HH:mm:ss}</green> {level} "
           "<blue>{message}</blue>", level="DEBUG")
logger.add('data/Журнал_обработанных_данных.log',
           format=" {time:HH:mm:ss} {level} {message}",
           level="DEBUG",
           rotation='50 MB', compression='zip', retention='7 days')


def create_data_folder():
    if not os.path.exists("data"):
        os.mkdir("data")


def get_driver_chrome():
    caps = DesiredCapabilities().CHROME
    caps["pageLoadStrategy"] = 'eager'
    options = webdriver.ChromeOptions()
    options.add_experimental_option('excludeSwitches', ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    options.headless = True
    driver = webdriver.Chrome(executable_path=PATH_TO_DRIVER,
                              options=options, desired_capabilities=caps)
    stealth(driver, user_agent=useragent, languages=["en-US", 'en'],
            vendor="Google Inc.", platform="Win 32",
            webgl_vendor="Intel Inc.", renderer="Intel Iris OpenGL Engine",
            fix_hairline=True)
    return driver


def ask_user_questions():
    try:
        answer = input(sub.questions)
        print()
        if answer.lower() == 'все':
            city, start_value = sub.city_list, 0
            sub.create_file_for_links()
            sub.сreate_excel_tables()
            request, num_handled = input(sub.question_3), 0
        elif answer == 'продолжить сбор ссылок':
            num_saved = int(input(sub.question_4))
            city = sub.city_list[num_saved:]
            start_value = input(sub.question_2)
            request, num_handled = input(sub.question_3), 0
        elif answer == 'продолжить обработку ссылок':
            num_handled = input(sub.question_5)
            city, start_value, request = {}, 0, None
            logger.info('Обработка началась...')
        elif answer.lower() == 'один город':
            name_city = input('Введите название города\n')
            city, start_value, num_handled = [(1, name_city)], 0, 0
            request = input(sub.question_3)
            sub.сreate_excel_tables()
            sub.create_file_for_links()
        return city, request, int(start_value), int(num_handled)
    except Exception:
        logger.warning('Похоже опечатка!!!, попробуйте ввести все'
                       ' данные еще раз')


def input_city_and_request(driver, curent_city, request):
    try:
        driver.get(URL)
        sleep(1)
        driver.find_element(By.TAG_NAME, "input").send_keys(curent_city)
        driver.find_element(By.TAG_NAME, "input").send_keys(" " + request)
        driver.find_element(By.TAG_NAME, "button").click()
        sleep(1.5)
    except Exception as ex:
        logger.error('ошибка при вводе поисковый данных', ex)
        sleep(10)
        input_city_and_request(driver, curent_city, request)


def scroll_page_down_links(driver, actions):
    '''Srcoll down for all main links seeing'''
    try:
        num_of_pushing_page_down = 19
        clickable_element = (driver.find_element(By.CLASS_NAME,
                             "search-list-view__content")
                             .find_element(By.TAG_NAME, "div"))
        actions.click(clickable_element).perform()
        page_scrolling = driver.find_element(By.TAG_NAME, "body")
        while True:  # srcoll action
            num_links_before_scroll = len(driver.find_elements(By.CLASS_NAME,
                                          "search-snippet-view"))
            for i in range(num_of_pushing_page_down):
                page_scrolling.send_keys(Keys.PAGE_DOWN)
                sleep(0.1)
            num_links_after_scroll = len(driver.find_elements(By.CLASS_NAME,
                                         "search-snippet-view"))
            if num_links_after_scroll > num_links_before_scroll:  # comparison
                continue
            else:
                break
    except Exception as ex:
        logger.error("Ошибка при скролинге страницы", ex)


def save_all_links(source, num_links):
    try:
        temp_list_for_links = []
        soup = BeautifulSoup(source, "lxml")
        all_links = (soup.find("ul", class_="search-list-view__list")
                         .find_all("a", class_="search-snippet-view__"
                                   "link-overlay _focusable"))
        for i in all_links:
            text = i.get("href")
            curent_link = f"{(num := num_links())} https://yandex.ru{text}"
            temp_list_for_links.append(curent_link)
        sub.save_links_in_txt(temp_list_for_links)
        return num, len(temp_list_for_links)
    except Exception as ex:
        logger.error('Ошибка при сохранении ссылок в файл', ex)


def get_all_links(current_city, request, num_links):
    try:
        logger.info(f'Идет сбор ссылок в городе {current_city[0]}'
                    f' {current_city[1]}')
        driver = get_driver_chrome()
        actions = ActionChains(driver)
        input_city_and_request(driver, current_city[1], request)
        scroll_page_down_links(driver, actions)
        num, sum_links = save_all_links(driver.page_source, num_links)
        logger.info(f'Для города {current_city[0]} {current_city[1]} '
                    f'собрано {sum_links} шт., собрано всего {num} шт.')
    except Exception as ex:
        logger.error(ex)
    finally:
        driver.close()
        driver.quit()


def get_data_for_partner_table(link, partner_id) -> int:
    try:
        driver = get_driver_chrome()
        driver.get(link)
        sleep(1)
        soup = BeautifulSoup(driver.page_source, "lxml")
        empty_line = ''
        result_list = [
            partner_id,
            'компания',
            sub.get_name_of_partner(soup),
            empty_line,
            sub.get_phone_number(soup),
            empty_line,
            sub.get_address(soup),
            empty_line,
            empty_line,
            sub.get_messenger(soup, 'telegram'),
            sub.get_messenger(soup, 'whatsapp'),
            empty_line,
            sub.get_messenger(soup, 'vkontakte'),
            sub.get_website(soup),
            empty_line,
            sub.get_logo_link(soup),
            sub.get_coordinates(driver)[1],
            sub.get_coordinates(driver)[0],
            sub.get_photos_links(driver, link)
        ]
        sub.append_data_table_partners(result_list)
        return partner_id
    except Exception as ex:
        logger.warning(ex)
    finally:
        driver.close()
        driver.quit()


def get_data_for_reviews_table(link, partner_id):
    try:
        driver = get_driver_chrome()
        driver.get(f'{link}reviews')
        sleep(1)
        sub.scroll_page_down_reviews(driver)
        soup = BeautifulSoup(driver.page_source, "lxml")
        all_reviews = sub.get_all_reviews(soup)
        empty_line = ''
        wb = load_workbook("data/reviews.xlsx")
        ws = wb.active
        if all_reviews:
            for one_block in all_reviews:
                current_review = [
                    partner_id,
                    sub.get_name_client(one_block),
                    sub.get_rating(one_block),
                    sub.get_date_review(one_block),
                    sub.get_text_review(one_block),
                    empty_line
                ]
                ws.append(current_review)
    except Exception as ex:
        logger.warning(ex)
    finally:
        wb.save("data/reviews.xlsx")
        wb.close()
        driver.close()
        driver.quit()


def get_data_for_services_table(link, partner_id):
    try:
        driver = get_driver_chrome()
        driver.get(f'{link}prices')
        sleep(1)
        soup = BeautifulSoup(driver.page_source, "lxml")
        all_services = sub.get_all_serevices(soup)
        wb = load_workbook("data/services.xlsx")
        ws = wb.active
        if all_services:
            for one_sevice in all_services:
                current_service = [
                    partner_id,
                    sub.get_name_service(one_sevice),
                    sub.get_price(one_sevice)
                ]
                ws.append(current_service)
    except Exception as ex:
        logger.warning(ex)
    finally:
        wb.save("data/services.xlsx")
        wb.close()
        driver.close()
        driver.quit()


def process_all_links(num_handled):

    with open("data/links.txt", "r") as file:
        for row in file:
            link = row.strip().split()[1]
            partner_id = int(row.strip().split()[0])
            if num_handled >= partner_id:
                continue
            try:
                tread_1 = Thread(target=get_data_for_partner_table,
                                 args=(link, partner_id))
                tread_2 = Thread(target=get_data_for_reviews_table,
                                 args=(link, partner_id))
                tread_3 = Thread(target=get_data_for_services_table,
                                 args=(link, partner_id))
                tread_3.start()
                tread_1.start()
                tread_2.start()
                tread_3.join()
                tread_1.join()
                tread_2.join()
                logger.info(f'Идет обработка ссылок, {partner_id}'
                            ' сохранена')
            except Exception as ex:
                logger.warning(ex)
    logger.info(f'Обработка ссылок окончена! Всего {partner_id} ссылок')


@logger.catch
def main():

    try:
        create_data_folder()
        city, request, start_value, num_handled = ask_user_questions()
        num_links = sub.counter(start_value)
        for current_city in city:
            get_all_links(current_city, request, num_links)
        process_all_links(num_handled)
    except Exception as ex:
        logger.error(ex)


if __name__ == '__main__':
    main()
