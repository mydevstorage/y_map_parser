from time import sleep
from loguru import logger
from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains

# DATA_FOLDER = '/home/roman/real_python/web_parsing/yandex_map_parser'
# DATA_FOLDER = 'yandex_map_parser'

all_cities = ['Москва район Академический', 'Москва район Алексеевский',
              'Москва район Алтуфьевский', 'Москва район Арбат',
              'Москва район Аэропорт', 'Москва район Бабушкинский',
              'Москва район Басманный', 'Москва район Беговой',
              'Москва район Бескудниковский', 'Москва район Бибирево',
              'Москва район Бирюлёво Восточное',
              'Москва район Бирюлёво Западное',
              'Москва район Богородское', 'Москва район Братеево',
              'Москва район Бутырский', 'Москва район Вешняки',
              'Москва район Внуково', 'Москва район Войковский',
              'Москва район Восточное Дегунино',
              'Москва район Восточное Измайлово',
              'Москва район Восточный', 'Москва район Выхино-Жулебино',
              'Москва район Гагаринский', 'Москва район Головинский',
              'Москва район Гольяново', 'Москва район Даниловский',
              'Москва район Дмитровский', 'Москва район Донской',
              'Москва район Дорогомилово', 'Москва район Замоскворечье',
              'Москва район Западное Дегунино',
              'Москва район Зюзино', 'Москва район Зябликово',
              'Москва район Ивановский', 'Москва район Измайлово',
              'Москва район Капотня', 'Москва район Коньково',
              'Москва район Коптево', 'Москва район Косино-Ухтомское',
              'Москва район Котловка', 'Москва район Красносельский',
              'Москва район Крылатское', 'Москва район Крюково',
              'Москва район Кузьминки', 'Москва район Кунцево',
              'Москва район Куркино', 'Москва район Левобережный',
              'Москва район Лефортово', 'Москва район Лианозово',
              'Москва район Ломоносовский', 'Москва район Лосиноостровский',
              'Москва район Люблино', 'Москва район Марфино',
              'Москва район Марьина Роща',
              'Москва район Марьино', 'Москва район Матушкино',
              'Москва район Метрогородок', 'Москва район Мещанский',
              'Москва район Митино', 'Москва район Можайский',
              'Москва район Молжаниновский',
              'Москва район Москворечье-Сабурово',
              'Москва район Нагатино-Садовники',
              'Москва район Нагатинский Затон', 'Москва район Нагорный'
              'Москва район Нижегородский', 'Москва район Новогиреево',
              'Москва район Новокосино', 'Москва район Ново-Переделкино',
              'Москва район Обручевский',
              'Москва район Орехово-Борисово южное',
              'Москва район Орехово-Борисово северное',
              'Москва район Останкинский',
              'Москва район Отрадное', 'Москва район Очаково-Матвеевское',
              'Москва район Перово', 'Москва район Печатники',
              'Москва район Покровское-Стрешнево',
              'Москва район Преображенское',
              'Москва район Пресненский',
              'Москва район проспект Вернадского', 'Москва район Раменки',
              'Москва район Ростокино', 'Москва район Рязанский',
              'Москва район Савёлки', 'Москва район Савёловский',
              'Москва район Свиблово', 'Москва район Северное бутово',
              'Москва район Северное измайлово',
              'Москва район Северное медведково',
              'Москва район Северное тушино',
              'Москва район Силино', 'Москва район Сокол',
              'Москва район Соколиная Гора',
              'Москва район Сокольники', 'Москва район Солнцево',
              'Москва район Старое Крюково',
              'Москва район Строгино', 'Москва район Таганский',
              'Москва район Тверской', 'Москва район Текстильщики',
              'Москва район Тёплый Стан',
              'Москва район Тимирязевский', 'Москва район Тропарёво-Никулино',
              'Москва район Филёвский Парк',
              'Москва район Фили-Давыдково', 'Москва район Хамовники',
              'Москва район Ховрино', 'Москва район Хорошёво-Мнёвники',
              'Москва район Хорошёвский', 'Москва район Царицыно',
              'Москва район Черёмушки', 'Москва район Чертаново северное',
              'Москва район Чертаново центральное',
              'Москва район Чертаново южное',
              'Москва район Щукино',
              'Москва район Южное Бутово',
              'Москва район Южное Медведково',
              'Москва район Южное Тушино',
              'Москва район Южнопортовый', 'Москва район Якиманка',
              'Москва район Ярославский', 'Москва район Ясенево',
              'Москва район Некрасовка'
              'Санкт-Петербург район Адмиралтейский',
              'Санкт-Петербург район Василеостровский',
              'Санкт-Петербург район Выборгский',
              'Санкт-Петербург район Калининский',
              'Санкт-Петербург район Кировский',
              'Санкт-Петербург район Колпинский',
              'Санкт-Петербург район Красногвардейский',
              'Санкт-Петербург район Красносельский',
              'Санкт-Петербург район Кронштадтский',
              'Санкт-Петербург район Курортный',
              'Санкт-Петербург район Московский',
              'Санкт-Петербург район Невский',
              'Санкт-Петербург район Петроградский',
              'Санкт-Петербург район Петродворцовый',
              'Санкт-Петербург район Приморский',
              'Санкт-Петербург район Пушкинский',
              'Санкт-Петербург район Фрунзенский',
              'Санкт-Петербург район Центральный',
              'Калуга', 'Брянск', 'Майкоп', 'Новосибирск',
              'Казань', 'Нижний Новгород', 'Челябинск', 'Красноярск', 'Самара',
              'Уфа', 'Ростов-на-Дону', 'Омск', 'Краснодар', 'Воронеж', 'Пермь',
              'Волгоград', 'Саратов', 'Тюмень', 'Тольятти', 'Барнаул',
              'Ижевск', 'Махачкала', 'Хабаровск', 'Ульяновск', 'Иркутск',
              'Владивосток', 'Ярославль', 'Кемерово', 'Томск', 'Тула',
              'Набережные Челны', 'Севастополь', 'Владимир', 'Грозный',
              'Ставрополь', 'Оренбург', 'Новокузнецк', 'Рязань', 'Балашиха',
              'Пенза', 'Чебоксары', 'Липецк', 'Калининград', 'Астрахань',
              'Киров', 'Сочи', 'Курск', 'Улан-Удэ', 'Тверь', 'Рыбинск',
              'Магнитогорск', 'Сургут', 'Иваново', 'Якутск',
              'Симферополь', 'Белгород', 'Нижний Тагил', 'Чита',
              'Волжский', 'Смоленск', 'Подольск', 'Саранск', 'Вологда',
              'Череповец', 'Орёл', 'Архангельск', 'Владикавказ',
              'Йошкар-Ола', 'Стерлитамак', 'Мурманск', 'Кострома',
              'Тамбов', 'Химки', 'Мытищи', 'Нальчик', 'Таганрог', 'Нижнекамск',
              'Благовещенск', 'Комсомольск-на-Амуре', 'Петрозаводск',
              'Шахты', 'Энгельс', 'Великий Новгород', 'Люберцы',
              'Братск', 'Старый Оскол', 'Ангарск', 'Сыктывкар', 'Дзержинск',
              'Псков', 'Орск', 'Красногорск', 'Армавир', 'Абакан', 'Балаково',
              'Бийск', 'Южно-Сахалинск', 'Одинцово', 'Уссурийск',
              'Норильск', 'Волгодонск', 'Сызрань', 'Петропавловск-Камчатский',
              'Новочеркасск', 'Альметьевск', 'Златоуст', 'Северодвинск',
              'Хасавюрт', 'Керчь', 'Домодедово', 'Салават', 'Миасс',
              'Копейск', 'Пятигорск', 'Электросталь', 'Находка',
              'Березники', 'Коломна', 'Щёлково', 'Серпухов', 'Ковров',
              'Кисловодск', 'Батайск', 'Рубцовск', 'Обнинск', 'Кызыл',
              'Нефтеюганск', 'Назрань', 'Каспийск', 'Долгопрудный',
              'Новомосковск', 'Ессентуки', 'Невинномысск', 'Октябрьский',
              'Раменское', 'Первоуральск', 'Михайловск', 'Реутов', 'Черкесск',
              'Жуковский', 'Димитровград', 'Пушкино', 'Артём', 'Камышин',
              'Евпатория', 'Муром', 'Ханты-Мансийск', 'Новый Уренгой',
              'Арзамас', 'Ногинск', 'Новошахтинск', 'Бердск', 'Элиста',
              'Северск', 'Новочебоксарск', 'Дербент', 'Нефтекамск',
              'Орехово-Зуево', 'Каменск-Уральский', 'Новороссийск',
              'Нижневартовск', 'Курган', 'Королёв', 'Прокопьевск',
              'Екатеринбург',
              'Ачинск', 'Тобольск', 'Ноябрьск', 'Видное', 'Сергиев Посад']

city_list = [(number, city) for number, city in enumerate(all_cities, 1)]

questions = ("Введите:\n"
             "а) <Один город>, для обработки одного города,\n"
             "б) <Все>' - для анализа всех городов России "
             "с населением более 100 тысяч человек...\n"
             "в) <продолжить сбор ссылок> - eсли произошло превывание "
             "программы при сборе ссылок\n"
             "г) <продолжить обработку ссылок> - eсли  произошло "
             "прерывание при обработке ссылок\n")

question_2 = ('Так же введите количество сохраненных ссылок'
              ' до прерывания (см. Журнал)\n')

question_3 = ("Введите сферу, которую хотите обработать/продолжить..\n")

question_4 = "номер последнего обработанного города(см. Журнал)\n"

question_5 = ("Введите количество обработанных ссылок (см.Журнал)"
              " для продолжения обработки с того же места\n")

months = {'января': '01', 'февраля': '02', 'марта': '03',
          'апреля': '04', 'мая': '05', 'июня': '06', 'июля': '07',
          'августа': '08', 'сентября': '09', 'октября': '10',
          'ноября': '11', 'декабря': '12'}


def counter(start=0):  # Cчетчик для компаний
    def number_plus():
        nonlocal start
        start += 1
        return start
    return number_plus


def сreate_excel_tables():
    headers = ['id_компании', 'Тип партнёра (мастер/компания)',
               'Имя', 'Фамилия', 'Телефон', 'Email', 'Адрес', 'О себе',
               'instagram', 'telegram', 'whatsapp', 'youtube', 'vk',
               'Личный сайт', 'odnoklassniki', 'Логотип', 'Широта',
               'Долгота', 'URL фотографий рабочего места, через |']
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    wb.save("data/partners.xlsx")
    wb.close()
    headers = ['Id мастера', 'Имя клиента', 'Рейтинг', 'Дата', 'Текст отзыва',
               'URL прикреплённых фотографий, через |']
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    wb.save("data/reviews.xlsx")
    wb.close()
    headers = ['Id партнёра', 'Название услуги',
               'Цена, руб', 'Длительность, минут']
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    wb.save("data/services.xlsx")
    wb.close()


def append_data_table_partners(saving_list: list):
    try:
        wb = load_workbook("data/partners.xlsx")
        ws = wb.active
        ws.append(saving_list)
    except Exception:
        logger.warning('Saving partner excel error')
    finally:
        wb.save("data/partners.xlsx")
        wb.close()


def get_name_of_partner(soup) -> str:
    try:
        name = (soup.find("div", class_="sticky-wrapper _position_top _header"
                " _border_auto _wide").find("div", class_="orgpage"
                "-header-view__header-wrapper")
                .find("h1", class_="orgpage-header-view__header")).text
        return name
    except Exception:
        return None


def get_phone_number(soup) -> str:
    try:
        tmp = (soup.find("div", class_="orgpage-header-view__contacts")
                   .find("div", class_="orgpage-phone"
                         "s-view__phone-number").text)
        number = (tmp.replace(') ', '').replace(' (', '').replace(' (', '')
                     .replace('-', ''))
        return number
    except Exception:
        return None


def get_address(soup) -> str:
    try:
        tmp = soup.find("a", class_="business"
                        "-contacts-view__address-link").text.split()
        address = (f'{tmp[-1]}, {" ".join(tmp[:-1])}').rstrip(',')
        return address
    except Exception:
        return None


def get_messenger(soup, name_messanger: str) -> str:
    try:
        all_messangers = soup.find_all("a", class_="button _view_secondary"
                                       "-gray _ui _size_medium _link")
        for tag in all_messangers:
            if name_messanger in tag.get("aria-label"):
                return tag.get('href').split('/')[-1]
    except Exception:
        return None


def get_website(soup):
    try:
        website = soup.find("span", class_="business-urls-view__text").text
        return website
    except Exception:
        return None


def get_logo_link(soup):
    try:
        logo = soup.find('img', class_='img-with-alt').get('src')
        return logo
    except Exception:
        return None


def get_coordinates(driver):
    try:
        coordinates = (driver.current_url.split("/")[7].replace("?ll=", "")
                       .replace("%2C", " ").split("&")[0].split())
        return coordinates[0], coordinates[1]
    except Exception:
        return None


def get_photos_links(driver, link):
    try:
        list_of_photo_links = []
        driver.get(f'{link}gallery')
        sleep(2)
        scroll_down_photo_page(driver)
        soup = BeautifulSoup(driver.page_source, "lxml")
        blocks_of_photo = soup.find_all("div", class_="photo-list__"
                                        "frame-wrapper")
        if blocks_of_photo:
            for item in blocks_of_photo:
                try:
                    link = item.find("img").get("src")
                    list_of_photo_links.append(link)
                except Exception:
                    continue
        return "|".join(list_of_photo_links)
    except Exception:
        return None


def get_all_reviews(soup):
    try:
        temp = 'business-reviews-card-view__reviews-container'
        all_reviews = (soup.find(class_=temp).find_all('div',
                       class_='business-reviews-card-view__review'))
        return all_reviews
    except Exception:
        return None


def get_name_client(soup):
    try:
        name = soup.find('span').text
        return name
    except Exception:
        return None


def get_rating(soup):
    try:
        rating = soup.find_all('span', class_='inline-image _loaded '
                               'business-rating-badge-view'
                               '__star _full _size_m')
        return len(rating)
    except Exception:
        return None


def get_date_review(soup):
    try:
        date = soup.find('span', class_='business-review-view__date').text
        temp = date.strip().split()
        a = '0' + temp[0] if len(temp[0]) == 1 else temp[0]
        if len(temp) == 3:
            return f'{temp[2]}-{months[temp[1]]}-{a}'
        elif len(temp) == 2:
            return f'2022-{months[temp[1]]}-{a}'
    except Exception:
        return None


def get_text_review(soup):
    try:
        text = soup.find('span', class_='business-review-view__body-text').text
        return text
    except Exception:
        return None


def get_all_serevices(soup):
    try:
        return soup.find_all('div', class_='business-'
                             'full-items-grouped-view__item _view_list')
    except Exception:
        return None


def get_name_service(soup):
    try:
        title = soup.find('div', class_='related-item-list-view__title').text
        return title
    except Exception:
        return None


def get_price(soup):
    try:
        price = soup.find('div', class_='related-item-list-view__price').text
        return price[:-1]
    except Exception:
        return None


def scroll_down_photo_page(driver):
    '''Scrolling down for all reviews seeing'''
    try:
        actions = ActionChains(driver)
        num_of_pushing_page_down = 20
        clickable_element = (driver.find_element(By.TAG_NAME,
                             "h1"))
        actions.click(clickable_element).perform()
        page_scrolling = driver.find_element(By.TAG_NAME, "body")
        while True:
            tmp = driver.find_elements(By.CLASS_NAME, "photo-wrapper__photo")
            for i in range(num_of_pushing_page_down):
                page_scrolling.send_keys(Keys.PAGE_DOWN)
                sleep(0.1)
            tmp2 = driver.find_elements(By.CLASS_NAME, "photo-wrapper__photo")
            if len(tmp) < len(tmp2):
                continue
            else:
                break
    except Exception:
        logger.warning('Scroll photo page')


def scroll_page_down_reviews(driver):
    '''Scrolling down for all reviews seeing'''
    try:
        actions = ActionChains(driver)
        num_of_pushing_page_down = 20
        clickable_element = (driver.find_element(By.TAG_NAME,
                             "h1"))
        actions.click(clickable_element).perform()
        page_scrolling = driver.find_element(By.TAG_NAME, "body")
        while True:
            tmp = len(driver.find_elements(By.CLASS_NAME, 'business-'
                      'reviews-card-view__review'))
            for i in range(num_of_pushing_page_down):
                page_scrolling.send_keys(Keys.PAGE_DOWN)
                sleep(0.1)
            tmp2 = len(driver.find_elements(By.CLASS_NAME, 'business-reviews-'
                       'card-view__review'))
            if tmp < tmp2:
                continue
            else:
                break
    except Exception:
        logger.warning('Scroll reviews page')


def create_file_for_links():
    with open("data/links.txt", 'w') as file:
        file.write('')


def save_links_in_txt(list):
    with open("data/links.txt", 'a') as file:
        for link in list:
            file.write(f'{link}\n')
